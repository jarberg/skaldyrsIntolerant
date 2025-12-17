import io

from openpyxl import load_workbook

def convert_excel_to_dict(excel_bytes):
    wb = load_workbook(io.BytesIO(excel_bytes))
    ws = wb.active
    #if wb.active.title == "Usage" or wb.active.title == "Acronis Total":
       #ws = wb.worksheets[1]
    # else:
      #  ws = wb.active

    # header row
    header_row = next(ws.iter_rows(values_only=True))
    headers = [cell for cell in header_row]

    row_dicts = []
    for row in list(ws.iter_rows(min_row=2, values_only=True)):
        row_dicts.append(
            dict(zip(headers, row))
        )
    return row_dicts

def convert_row_to_dict(catKey, row):
    ret = {
        "Category": catKey,
        "InvoicePeriodStart": row.get("Start Date", ""),
        "InvoicePeriodEnd": row.get("End Date", ""),
        "Item Name": row.get("Item Name", ""),
        "ItemNo": row.get("ItemNo", ""),
        "Quantity": row.get("Quantity",""),
        "Unit Price": row.get("Unit Price", ""),
        "Units": row.get("Units", "stk"),
        "Amount": row,
        "Currency": row.get("Currency", ""),
        "Customer Id": "",
        "Customer Name (from Excel)": "",
        "VAT (from Excel)": "",
        "Note": "No customer id column in billing file",
    }


    return ret

def get_id_keys(data_rows):
    success = True
    id_key = None
    vat_key = None
    name_key = None

    firstRow = next(iter(data_rows))
    if "Portal Customer Id" in firstRow:
        id_key = "Portal Customer Id"
        vat_key = "Portal Customer VAT"
        name_key = "Portal Customer Name"
    elif "Customer Id" in firstRow:
        # Acronis / Impossible Cloud style
        id_key = "Customer Id"
        vat_key = "Customer VAT" if "Customer VAT" in firstRow else None
        name_key = "Customer Name"
    else:
        success = False


    return id_key, vat_key, name_key, success