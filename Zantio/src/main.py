global DRY_RUN

from dotenv import load_dotenv

load_dotenv()

import io
import csv
import json
from pathlib import Path

from RESTclients.dataModels import (
    CustomerInvoice,
    CustomerInvoiceCategory,
    CustomerInvoiceCategoryLine_exclaimer,
    CustomerInvoiceCategoryLine_keepit,
    CustomerInvoice_Error,
)
from openpyxl import load_workbook

import util as util
from RESTclients import cloudfactory as cf, uniconta as uc


# --------------------------------------------------------------------
# Output directory (relative to this file: src/output)
# --------------------------------------------------------------------
BASE_DIR = Path(__file__).resolve().parent
OUTPUT_DIR = BASE_DIR / "output"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


def compute_line_total(line) -> float:
    """
    POTENTIEL hjælp: beløb = Quantity × UnitPrice, fallback til Amount.
    Pt. bruger vi konsekvent Amount til alle summeringer, da Amount er
    det økonomisk korrekte tal (tidsbaserede licenser m.m.).
    """
    try:
        qty = float(getattr(line, "Quantity", 0) or 0)
        price = float(getattr(line, "UnitPrice", 0) or 0)
        total = qty * price
        if total == 0 and getattr(line, "Amount", None) not in (None, "", 0):
            total = float(line.Amount or 0)
        return total
    except Exception:
        return 0.0


def generate_correct_product_line(catName, record):
    if catName == "Exclaimer":
        line = CustomerInvoiceCategoryLine_exclaimer(
            Amount=record.get("Amount", "Failed"),
            Currency=record.get("Currency", "Failed"),
            ItemName=record.get("Item Name", "Failed"),
            ItemNo=record.get("ItemNo", "Failed"),
            LicenseAgreementType=record.get("License Agreement Type", "Failed"),
            Offering=record.get("Offering", "Failed"),
            ProductFamily=record.get("Product Family", "Failed"),
            Quantity=record.get("Quantity", "Failed"),
            UnitPrice="",
        )
    elif catName in [
        "Keepit",
        "Impossible Cloud",
        "Acronis",
        "Dropbox",
        "Microsoft CSP (NCE)",
        "SPLA",
    ]:
        line = CustomerInvoiceCategoryLine_keepit(
            Amount=record.get("Amount", "Failed"),
            Currency=record.get("Currency", "Failed"),
            ItemName=record.get("Item Name", "Failed"),
            ItemNo=record.get("ItemNo", "Failed"),
            LicenseAgreementType=record.get("License Agreement Type", "Failed"),
            Offering=record.get("Offering", "Failed"),
            ProductFamily=record.get("Product Family", "Failed"),
            Quantity=record.get("Quantity", "Failed"),
            UnitPrice="",
        )
    else:
        line = CustomerInvoiceCategoryLine_keepit(
            Amount=record.get("Amount", "Failed"),
            Currency=record.get("Currency", "Failed"),
            ItemName=record.get("Item Name", "Failed"),
            ItemNo=record.get("ItemNo", "Failed"),
            LicenseAgreementType=record.get("License Agreement Type", "Failed"),
            Offering=record.get("Offering", "Failed"),
            ProductFamily=record.get("Product Family", "Failed"),
            Quantity=record.get("Quantity", "Failed"),
            UnitPrice="",
        )
    return line


def print_invoices(invoiceCustomerdict):
    for invoice in invoiceCustomerdict.values():
        print(invoice.customer.name)
        for key in invoice.categories.keys():
            category = invoice.categories.get(key)
            print("   |_ " + key)
            for line in category.lines:
                print("         |_ itemName: " + line.ItemName)
                print("         |_ ProductFamily: " + line.ProductFamily)
                print("         |_ amount: " + str(line.Amount))
                print("         |_ quantity: " + str(line.Quantity))
                print("         |_ unitPrice: " + str(line.UnitPrice))
                print("         |")
        print("")


def export_failed_customers_csv(failedCustomerlist, output_path: Path) -> float:
    """
    CloudFactory-rækker der IKKE kunne matches til en CloudFactory-kunde.
    Beløb beregnes som Amount (CloudFactory-beløb).
    Returnerer totalbeløb.
    """
    rows = []
    total_failed = 0.0

    for customerid, customer_invoice in failedCustomerlist.items():
        reason = getattr(customer_invoice, "reason", "Unknown")

        total_amount = 0.0
        for category in customer_invoice.categories.values():
            for line in category.lines:
                try:
                    total_amount += float(line.Amount or 0)
                except Exception:
                    pass

        total_failed += total_amount

        customer_name = (
            customer_invoice.customer.name
            if customer_invoice.customer else "Unknown"
        )
        vat = (
            customer_invoice.customer.vatID
            if customer_invoice.customer else "Unknown"
        )

        rows.append({
            "Customer ID": customerid,
            "Customer Name": customer_name,
            "VAT": vat,
            "Reason": reason,
            "Total Amount (DKK)": total_amount,
        })

    if not rows:
        print("No failed CloudFactory customers – nothing to write to CSV.")
        return 0.0

    output_file = Path(output_path)
    with output_file.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=rows[0].keys())
        writer.writeheader()
        writer.writerows(rows)

    print(f"Failed customer CSV exported -> {output_file.resolve()}")
    return total_failed


def export_missing_debtors_csv(failed_invoices, output_path: Path) -> float:
    """
    Kunder hvor vi IKKE fandt en debitor i Uniconta.
    Beløb beregnes som Amount.
    Returnerer totalbeløb.
    """
    rows = []
    total_failed = 0.0

    for inv in failed_invoices:
        cust = inv.customer
        if not cust:
            continue

        total_amount = 0.0
        for category in inv.categories.values():
            for line in category.lines:
                try:
                    total_amount += float(line.Amount or 0)
                except Exception:
                    pass

        total_failed += total_amount

        rows.append({
            "Customer ID": cust.id,
            "Customer Name": cust.name,
            "VAT": cust.vatID,
            "Country": cust.countryCode,
            "Reason": "No matching debtor in Uniconta (find_deptor returned none)",
            "Total Amount (DKK)": total_amount,
        })

    if not rows:
        print("No missing debtors – all CustomerInvoices found a debtor in Uniconta.")
        return 0.0

    output_file = Path(output_path)
    with output_file.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=rows[0].keys())
        writer.writeheader()
        writer.writerows(rows)

    print(f"Missing debtors CSV exported -> {output_file.resolve()}")
    return total_failed


def export_success_invoices_csv(success_rows, output_path: Path) -> float:
    """
    Kunder hvor der blev oprettet en ordre i Uniconta.
    Rækkerne skal allerede indeholde 'Total Amount (DKK)' baseret på Amount.
    """
    if not success_rows:
        print("No successful invoices to export.")
        return 0.0

    total_success = sum(r.get("Total Amount (DKK)", 0.0) for r in success_rows)

    output_file = Path(output_path)
    with output_file.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=success_rows[0].keys())
        writer.writeheader()
        writer.writerows(success_rows)

    print(f"Success invoices CSV exported -> {output_file.resolve()}")
    return total_success


def export_no_customerid_csv(no_id_rows, output_path: Path) -> float:
    """
    Linjer i billing-Excel uden Customer Id (enten mangler id-kolonne
    helt eller feltet er tomt). Beløb = Amount.
    Returnerer totalbeløbet.
    """
    if not no_id_rows:
        print("No 'no customer id' rows – nothing to write to CSV.")
        return 0.0

    total = 0.0
    for r in no_id_rows:
        try:
            total += float(r.get("Amount", 0.0) or 0.0)
        except Exception:
            pass

    output_file = Path(output_path)
    with output_file.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=no_id_rows[0].keys())
        writer.writeheader()
        writer.writerows(no_id_rows)

    print(f"No-customer-id CSV exported -> {output_file.resolve()}")
    return total


def main(*args, **kwargs):
    # DRY_RUN-flag (pt. ikke brugt til at slå Uniconta-kald fra)
    DRY_RUN = kwargs.get("DRY_RUN", "True").lower() == "true"

    util.get_latest_date()

    cloudFac_client = cf.CloudFactoryClient()
    uniconta_client = uc.UnicontaAdapter()

    print("Fetching customers from CloudFactory...")
    uniconta_client.customerDataBase = cloudFac_client.list_customers()
    print(f"Found {len(uniconta_client.customerDataBase)} customers.")

    invoices = cloudFac_client.fetch_latest_invoices()
    invoiceCustomerdict: dict[str, CustomerInvoice] = {}
    failedCustomerlist: dict[str, CustomerInvoice_Error] = {}

    foundCatKeyDict = set()

    # Reconciliation totals
    total_amount_all = 0.0              # ALLE linjer med kunde-id (CloudFactory Amount)
    total_amount_success = 0.0          # Bogført mod Uniconta (Amount)
    total_amount_failed = 0.0           # Ingen debitor i Uniconta (Amount)
    total_amount_no_customerid = 0.0    # Linjer uden Customer Id ELLER uden id-kolonne

    # For success CSV
    success_rows = []
    # For no-customer-id CSV
    no_customerid_rows = []

    for invoice in invoices:
        for catKey in invoice.categories.keys():
            foundCatKeyDict.add(catKey)

            # load excel bytes
            excel_bytes = cloudFac_client.fetch_billing_excel(
                invoice.categories.get(catKey).excelLink
            )
            wb = load_workbook(io.BytesIO(excel_bytes))
            ws = wb.active

            # header row
            header_row = next(ws.iter_rows(values_only=True))
            headers = [cell for cell in header_row]

            # --- Determine which columns hold customer id / vat / name ---
            if "Portal Customer Id" in headers:
                id_key = "Portal Customer Id"
                vat_key = "Portal Customer VAT"
                name_key = "Portal Customer Name"
            elif "Customer Id" in headers:
                # Acronis / Impossible Cloud style
                id_key = "Customer Id"
                vat_key = "Customer VAT" if "Customer VAT" in headers else None
                name_key = "Customer Name"
            else:
                # 🔹 Ingen brugbar id-kolonne overhovedet (fx visse Acronis-filer)
                data_rows = list(ws.iter_rows(min_row=2, values_only=True))
                print(
                    f"[DEBUG] Category '{catKey}' has NO usable customer id column. "
                    f"Headers: {headers} | rows={len(data_rows)}"
                )
                for row in data_rows:
                    record = dict(zip(headers, row))
                    amount_raw = record.get("Amount", 0) or 0
                    try:
                        amount_val = float(amount_raw)
                    except (TypeError, ValueError):
                        amount_val = 0.0
                    total_amount_no_customerid += amount_val

                    no_customerid_rows.append(
                        {
                            "Category": catKey,
                            "InvoicePeriodStart": record.get("Start Date", ""),
                            "InvoicePeriodEnd": record.get("End Date", ""),
                            "Item Name": record.get("Item Name", ""),
                            "ItemNo": record.get("ItemNo", ""),
                            "Quantity": record.get("Quantity", ""),
                            "Unit Price": record.get("Unit Price", ""),
                            "Amount": amount_val,
                            "Currency": record.get("Currency", ""),
                            "Customer Id": "",
                            "Customer Name (from Excel)": "",
                            "VAT (from Excel)": "",
                            "Note": "No customer id column in billing file",
                        }
                    )
                # Vi kan ikke bygge CustomerInvoice uden id → videre til næste kategori
                continue

            # collect all data rows so we can both count and iterate
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            print(
                f"[DEBUG] Category '{catKey}' contains {len(data_rows)} rows in Excel."
            )

            previousCustomerid = None

            for row in data_rows:
                record = dict(zip(headers, row))

                # Beløb fra CloudFactory (Amount-kolonnen i Excel)
                amount_raw = record.get("Amount", 0) or 0
                try:
                    amount_val = float(amount_raw)
                except (TypeError, ValueError):
                    amount_val = 0.0

                raw_id = record.get(id_key)

                if not raw_id:
                    # 🔹 Ingen Customer Id i denne række → kan ikke tildele til kunde
                    total_amount_no_customerid += amount_val

                    no_customerid_rows.append(
                        {
                            "Category": catKey,
                            "InvoicePeriodStart": record.get("Start Date", ""),
                            "InvoicePeriodEnd": record.get("End Date", ""),
                            "Item Name": record.get("Item Name", ""),
                            "ItemNo": record.get("ItemNo", ""),
                            "Quantity": record.get("Quantity", ""),
                            "Unit Price": record.get("Unit Price", ""),
                            "Amount": amount_val,
                            "Currency": record.get("Currency", ""),
                            "Customer Id": "",
                            "Customer Name (from Excel)": record.get(name_key, "") if 'name_key' in locals() else "",
                            "VAT (from Excel)": record.get(vat_key, "") if 'vat_key' in locals() and vat_key else "",
                            "Note": "Row has id column but Customer Id is empty",
                        }
                    )
                    continue  # hop resten af kundelogikken over

                # Herfra ved vi at der er et Customer Id
                customerid = (
                    str(raw_id).replace("{", "").replace("}", "").lower()
                )

                vatID = record.get(vat_key, "NULL") if vat_key else "NULL"
                name = record.get(name_key, "NULL")

                # Build / reuse the CustomerInvoice / CustomerInvoice_Error
                customerinvoice = generate_customer_invoice(
                    previousCustomerid,
                    customerid,
                    vatID,
                    name,
                    invoiceCustomerdict,
                    record,
                    uniconta_client,
                    failedCustomerlist,
                )

                # Create / get the category on that invoice
                category_name = catKey
                category = customerinvoice.categories.get(category_name, None)
                if not category:
                    category = CustomerInvoiceCategory(
                        name=category_name,
                        lines=[],
                    )
                    customerinvoice.categories[category_name] = category

                # Create the line from CloudFactory record
                line = generate_correct_product_line(category_name, record)
                category.lines.append(line)
                previousCustomerid = customerid

                # 🔹 Per-customer CloudFactory total (kun rækker med kunde-id)
                total_amount_all += amount_val

    failedList: list[CustomerInvoice] = []

    # Actually create invoices in Uniconta
    for customerInvoice in invoiceCustomerdict.values():
        before = len(failedList)
        uniconta_client.create_invoice(customerInvoice, failedList)
        after = len(failedList)

        # Sum invoice amount: ren Amount (CloudFactory-beløb)
        inv_amount = 0.0
        for category in customerInvoice.categories.values():
            for line in category.lines:
                try:
                    inv_amount += float(line.Amount or 0)
                except (TypeError, ValueError):
                    pass

        if after == before:
            # No new failure added => this invoice was successfully posted
            total_amount_success += inv_amount

            cust = customerInvoice.customer
            success_rows.append(
                {
                    "Customer ID": cust.id,
                    "Customer Name": cust.name,
                    "VAT": cust.vatID,
                    "Country": cust.countryCode,
                    "Total Amount (DKK)": inv_amount,
                }
            )
        else:
            # This invoice ended up in failedList
            total_amount_failed += inv_amount

    # ------------------------------------------------------------------
    # Export CSVs for Streamlit / reporting
    # ------------------------------------------------------------------
    # 1) Successful invoices
    success_csv_path = OUTPUT_DIR / "success_invoices.csv"
    total_success_csv = export_success_invoices_csv(success_rows, success_csv_path)

    # 2) CloudFactory mapping failures (failedCustomerlist)
    failed_cf_csv_path = OUTPUT_DIR / "failed_customers.csv"
    total_failed_cf = export_failed_customers_csv(failedCustomerlist, failed_cf_csv_path)

    # 3) Uniconta debtor failures (failedList)
    failed_debtors_csv_path = OUTPUT_DIR / "failed_debtors_uniconta.csv"
    total_failed_debtors_csv = export_missing_debtors_csv(failedList, failed_debtors_csv_path)

    # 4) No customer id lines
    no_customerid_csv_path = OUTPUT_DIR / "no_customerid_lines.csv"
    total_no_customerid_csv = export_no_customerid_csv(no_customerid_rows, no_customerid_csv_path)

    # Debug: which categories did we see at all?
    print("\n=== CloudFactory billing categories discovered ===")
    print(foundCatKeyDict)
    print("=================================================\n")

    # --- Reconciliation summary (console) ---
    print("=== RECONCILIATION SUMMARY (ex. VAT) ===")
    print(f"Total CloudFactory per-customer amount processed         : {total_amount_all:,.2f} DKK")
    print(f"  → Mapped to existing Uniconta debtors                  : {total_amount_success:,.2f} DKK")
    print(f"  → No matching debtor in Uniconta (skipped)             : {total_amount_failed:,.2f} DKK")
    print(f"  → No matching CloudFactory customer (failed mapping)   : {total_failed_cf:,.2f} DKK")
    print(f"  → Lines with NO customer id in billing file            : {total_amount_no_customerid:,.2f} DKK")
    print(
        "Check (success + failed_debtors + failed_cf_customers)   : "
        f"{(total_amount_success + total_amount_failed + total_failed_cf):,.2f} DKK"
    )
    print(
        "Invoice header total (per-customer + no-id)              : "
        f"{(total_amount_all + total_amount_no_customerid):,.2f} DKK"
    )
    print("=========================================\n")

    for failed in failedList:
        print(failed)

    # ------------------------------------------------------------------
    # Write reconciliation JSON for Streamlit
    # ------------------------------------------------------------------
    reconciliation = {
        "total_cloudfactory_amount": round(total_amount_all, 2),
        "total_success_amount": round(total_amount_success, 2),
        "total_failed_debtors_amount": round(total_amount_failed, 2),
        "total_failed_cloudfactory_customers_amount": round(total_failed_cf, 2),
        "total_no_customerid_amount": round(total_amount_no_customerid, 2),
        "success_csv": str(success_csv_path),
        "failed_customers_csv": str(failed_cf_csv_path),
        "failed_debtors_csv": str(failed_debtors_csv_path),
        "no_customerid_csv": str(no_customerid_csv_path),
        "calculation_notes_da": {
            "cloudfactory_total": (
                "CloudFactory totalen (per kunde) er baseret på feltet 'Amount' i "
                "billing-Excel (sum pr. kunde med gyldigt kunde-id, ekskl. moms)."
            ),
            "booked_to_debtors": (
                "'Bogført mod Uniconta-debitorer' er summen af CloudFactory 'Amount' "
                "for alle fakturaer, hvor der blev oprettet en ordre i Uniconta."
            ),
            "missing_debtors": (
                "'Ingen debitor i Uniconta' er summen af CloudFactory 'Amount' "
                "for fakturaer, hvor vi ikke kunne finde en debitor i Uniconta."
            ),
            "missing_cf_customers": (
                "'Ingen kundematch i CloudFactory' er summen af CloudFactory 'Amount' "
                "for linjer, hvor vi ikke kunne matche en CloudFactory-kunde (kunde-fejl)."
            ),
            "no_customerid": (
                "‘Linjer uden Customer Id’ er beløb fra billing-Excel (Amount), "
                "hvor der enten ikke findes nogen kunde-id-kolonne overhovedet, "
                "eller hvor Customer Id-feltet er tomt. De kan ikke kobles til en debitor, "
                "men er vigtige for den samlede fakturasum (fx enkelte Acronis-linjer)."
            ),
        },
    }

    recon_path = OUTPUT_DIR / "reconciliation_summary.json"
    with recon_path.open("w", encoding="utf-8") as f:
        json.dump(reconciliation, f, indent=2, ensure_ascii=False)

    print(f"Reconciliation JSON exported -> {recon_path.resolve()}")


def generate_customer_invoice(
    previousCustomerid,
    customerid,
    vatID,
    name,
    invoiceCustomerdict,
    record,
    uniconta_adapter,
    failedCustomerlist,
):

    if (
        previousCustomerid != customerid
        and (customerid not in invoiceCustomerdict.keys())
        and (customerid not in failedCustomerlist.keys())
    ):
        potential_clients = list(
            (
                x
                for x in uniconta_adapter.customerDataBase
                if (x.id.lower() == customerid.lower())
            )
        )
        match len(potential_clients):
            case 0:
                customerinvoice = CustomerInvoice_Error(
                    customer=None,
                    reason="No match found for customer with name: "
                    + name
                    + " with vatid: "
                    + vatID,
                    categories={},
                )
                failedCustomerlist[customerid] = customerinvoice
            case 1:
                customerinvoice = CustomerInvoice(
                    customer=potential_clients[0],
                    period_start=record.get("Start Date", "ERROR"),
                    period_end=record.get("End Date", "ERROR"),
                    categories={},
                )
                invoiceCustomerdict[customerid] = customerinvoice
            case _:
                customerinvoice = CustomerInvoice_Error(
                    customer=None,
                    reason="to many matches found for customer with name: "
                    + name
                    + " with vatid: "
                    + vatID,
                    categories={},
                )
                failedCustomerlist[customerid] = customerinvoice
    else:
        if customerid in failedCustomerlist.keys():
            customerinvoice = failedCustomerlist.get(customerid)
        else:
            customerinvoice = invoiceCustomerdict.get(customerid)

    if not customerinvoice:
        customerinvoice = CustomerInvoice_Error(
            customer=None,
            reason="Base catched error. Customerid: "
            + customerid
            + " VatID: "
            + vatID
            + " Name: "
            + name
            + "",
            categories={},
        )
        failedCustomerlist[customerid] = customerinvoice

    return customerinvoice


if __name__ == "__main__":
    main()
